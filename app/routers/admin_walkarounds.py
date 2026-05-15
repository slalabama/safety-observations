from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends, Request
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List, Optional
import io
import os
from pathlib import Path

from app.database import SessionLocal
from app.models import WalkaroundForm, WalkaroundSection, WalkaroundQuestion
from app.routers.admin_auth import get_current_employee
from app.utils.pdf_ocr import extract_text_from_pdf

router = APIRouter(prefix="/walkarounds", tags=["Walkarounds"])

class QuestionCreate(BaseModel):
    text: str
    question_type: str = "pass_fail"
    order: int = 0

class SectionCreate(BaseModel):
    name: str
    order: int = 0

class WalkaroundFormCreate(BaseModel):
    name: str
    description: str = None

class OCRResult(BaseModel):
    sections: List[dict]

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Excel form parser
# ---------------------------------------------------------------------------
# Expected layout (case-insensitive headers):
#   Row 1: Section | Question | Type
#   Subsequent rows: one question per row. The Section column may repeat the
#   section name on every row OR be filled only on the first row of each
#   section (Excel-style merged-look). Both work.
#
# Type values understood (anything else falls back to pass_fail):
#   pass_fail, yes_no, text, date, number  (and common variants like
#   "Pass/Fail", "Y/N", "Numeric", etc.)
# ---------------------------------------------------------------------------

_TYPE_NORMALIZE = {
    "pass_fail": "pass_fail", "pass/fail": "pass_fail", "pass-fail": "pass_fail",
    "passfail": "pass_fail", "p/f": "pass_fail",
    "yes_no_na": "yes_no_na", "yes/no/na": "yes_no_na", "yes/no/n/a": "yes_no_na",
    "yes-no-na": "yes_no_na", "yesnona": "yes_no_na", "y/n/na": "yes_no_na",
    "y/n/n/a": "yes_no_na", "ynna": "yes_no_na",
    "yes_no": "yes_no", "yes/no": "yes_no", "yes-no": "yes_no",
    "yesno": "yes_no", "y/n": "yes_no",
    "text": "text", "string": "text", "freeform": "text", "comment": "text",
    "date": "date",
    "number": "number", "numeric": "number", "num": "number", "integer": "number",
}


def _norm_header(s):
    return (str(s).strip().lower() if s is not None else "")


def parse_excel_form(file_bytes: bytes) -> List[dict]:
    """Parse an Excel workbook into [{name, questions:[{text, question_type}]}]."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = [r for r in ws.iter_rows(values_only=True)]
    if not rows:
        return []

    # Locate columns by header name; fall back to positional if no match
    headers = [_norm_header(c) for c in rows[0]]
    def col(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return -1

    section_col  = col("section", "section name", "sections", "category")
    question_col = col("question", "question text", "text", "item", "inspection item", "check")
    type_col     = col("type", "question type", "format", "response type", "answer type")

    if question_col < 0:
        # No recognizable headers — assume positional: A=Section, B=Question, C=Type
        section_col, question_col, type_col = 0, 1, 2

    sections: List[dict] = []
    current = None

    for row in rows[1:]:
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue

        def cell(i):
            if i < 0 or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        sec_val = cell(section_col)
        q_val   = cell(question_col)
        t_val   = cell(type_col).lower()

        # Start a new section when we see a new section name
        if sec_val and (current is None or current["name"] != sec_val):
            current = {"name": sec_val, "questions": []}
            sections.append(current)

        if not q_val:
            continue

        if current is None:
            current = {"name": "General", "questions": []}
            sections.append(current)

        question_type = _TYPE_NORMALIZE.get(t_val.replace(" ", ""), None) \
                     or _TYPE_NORMALIZE.get(t_val, "pass_fail")

        current["questions"].append({
            "text": q_val,
            "question_type": question_type,
        })

    return sections

@router.post("/upload-pdf")
async def upload_pdf(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin = Depends(get_current_employee)
):
    """
    Upload PDF form and extract text for OCR.
    Returns structured data that admin can review/edit before creating form.
    """
    
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be PDF")
    
    # Save PDF temporarily
    upload_dir = Path("uploads/walkarounds")
    upload_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = upload_dir / file.filename
    
    with open(pdf_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    # Extract text
    try:
        extracted_data = extract_text_from_pdf(str(pdf_path))
        return OCRResult(sections=extracted_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF extraction failed: {str(e)}")

@router.post("/create-from-ocr")
def create_form_from_ocr(
    form_data: dict,
    db: Session = Depends(get_db),
    admin = Depends(get_current_employee)
):
    """
    Create walkaround form from OCR data (admin-reviewed).
    """
    
    form = WalkaroundForm(name=form_data.get("name"), description=form_data.get("description"))
    db.add(form)
    db.flush()
    
    for sec_idx, section_data in enumerate(form_data.get("sections", [])):
        section = WalkaroundSection(
            form_id=form.id,
            name=section_data.get("name"),
            order=sec_idx
        )
        db.add(section)
        db.flush()
        
        for q_idx, question_data in enumerate(section_data.get("questions", [])):
            question = WalkaroundQuestion(
                section_id=section.id,
                text=question_data.get("text"),
                question_type=question_data.get("question_type", "pass_fail"),
                order=q_idx
            )
            db.add(question)
    
    db.commit()
    db.refresh(form)
    
    return {"id": form.id, "name": form.name, "message": "Form created from OCR"}


@router.post("/upload")
async def upload_form(
    name: str = Form(...),
    description: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin = Depends(get_current_employee),
):
    """
    Upload a PDF or Excel form file and create the walkaround form with its
    sections and questions in one step. PDF goes through pdf_ocr; Excel goes
    through the local parse_excel_form helper.
    """
    filename = (file.filename or "").lower()
    contents = await file.read()

    if filename.endswith(".pdf"):
        # Reuse the existing PDF extractor; it expects a file path on disk
        upload_dir = Path("uploads/walkarounds")
        upload_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = upload_dir / file.filename
        with open(pdf_path, "wb") as fp:
            fp.write(contents)
        try:
            sections = extract_text_from_pdf(str(pdf_path))
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"PDF parsing failed: {e}")
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            sections = parse_excel_form(contents)
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl is not installed; add 'openpyxl' to requirements.txt and redeploy",
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Excel parsing failed: {e}")
    else:
        raise HTTPException(
            status_code=400,
            detail="File must be .pdf, .xlsx, or .xls",
        )

    if not sections:
        raise HTTPException(
            status_code=400,
            detail="No sections/questions were found in the file. Check the layout.",
        )

    # Create the form, sections, and questions in a single transaction
    form = WalkaroundForm(name=name, description=description)
    db.add(form)
    db.flush()

    total_questions = 0
    for sec_idx, section_data in enumerate(sections):
        section = WalkaroundSection(
            form_id=form.id,
            name=section_data.get("name") or f"Section {sec_idx + 1}",
            order=sec_idx,
        )
        db.add(section)
        db.flush()

        for q_idx, qd in enumerate(section_data.get("questions", [])):
            db.add(WalkaroundQuestion(
                section_id=section.id,
                text=qd.get("text", "").strip(),
                question_type=qd.get("question_type", "pass_fail"),
                order=q_idx,
            ))
            total_questions += 1

    db.commit()
    db.refresh(form)

    return {
        "id": form.id,
        "name": form.name,
        "section_count": len(sections),
        "question_count": total_questions,
        "message": f"Form created with {len(sections)} section(s) and {total_questions} question(s).",
    }


@router.delete("/{form_id}")
def delete_form(
    form_id: int,
    db: Session = Depends(get_db),
    admin = Depends(get_current_employee),
):
    """
    Delete a walkaround form.
    - If the form has submissions, soft-delete it (active=False) so submission
      history and PDF generation keep working.
    - Otherwise hard-delete the form and its sections/questions.
    """
    from app.models import WalkaroundSubmission

    form = db.query(WalkaroundForm).filter(WalkaroundForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")

    submission_count = (
        db.query(WalkaroundSubmission)
          .filter(WalkaroundSubmission.form_id == form_id)
          .count()
    )

    if submission_count > 0:
        form.active = False
        db.commit()
        return {
            "deleted": False,
            "soft_deleted": True,
            "submission_count": submission_count,
            "message": (
                f"Form deactivated. {submission_count} submission(s) preserved "
                f"so existing PDFs and history still work."
            ),
        }

    # No submissions — hard delete is safe. Sections and questions cascade.
    db.delete(form)
    db.commit()
    return {"deleted": True, "soft_deleted": False, "message": "Form deleted"}

@router.post("/")
def create_form(
    form_data: WalkaroundFormCreate,
    db: Session = Depends(get_db),
    admin = Depends(get_current_employee)
):
    """Create a new walkaround form manually."""
    form = WalkaroundForm(name=form_data.name, description=form_data.description)
    db.add(form)
    db.commit()
    db.refresh(form)
    
    return {"id": form.id, "name": form.name, "message": "Form created"}

@router.get("/")
def list_forms(
    db: Session = Depends(get_db),
    admin = Depends(get_current_employee)
):
    """List all walkaround forms."""
    forms = db.query(WalkaroundForm).all()
    return [
        {
            "id": f.id,
            "name": f.name,
            "description": f.description,
            "active": f.active,
            "section_count": len(f.sections)
        } for f in forms
    ]

@router.get("/{form_id}")
def get_form(
    form_id: int,
    db: Session = Depends(get_db),
    admin = Depends(get_current_employee)
):
    """Get a walkaround form with sections and questions."""
    form = db.query(WalkaroundForm).filter(WalkaroundForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    return {
        "id": form.id,
        "name": form.name,
        "description": form.description,
        "active": form.active,
        "sections": [
            {
                "id": s.id,
                "name": s.name,
                "order": s.order,
                "questions": [
                    {
                        "id": q.id,
                        "text": q.text,
                        "question_type": q.question_type,
                        "order": q.order
                    } for q in s.questions
                ]
            } for s in form.sections
        ]
    }

@router.post("/{form_id}/sections")
def add_section(
    form_id: int,
    section: SectionCreate,
    db: Session = Depends(get_db),
    admin = Depends(get_current_employee)
):
    """Add a section to a form."""
    form = db.query(WalkaroundForm).filter(WalkaroundForm.id == form_id).first()
    if not form:
        raise HTTPException(status_code=404, detail="Form not found")
    
    s = WalkaroundSection(form_id=form_id, name=section.name, order=section.order)
    db.add(s)
    db.commit()
    db.refresh(s)
    
    return {"id": s.id, "name": s.name, "order": s.order}

@router.post("/{form_id}/sections/{section_id}/questions")
def add_question(
    form_id: int,
    section_id: int,
    question: QuestionCreate,
    db: Session = Depends(get_db),
    admin = Depends(get_current_employee)
):
    """Add a question to a section."""
    section = db.query(WalkaroundSection).filter(
        WalkaroundSection.id == section_id,
        WalkaroundSection.form_id == form_id
    ).first()
    
    if not section:
        raise HTTPException(status_code=404, detail="Section not found")
    
    q = WalkaroundQuestion(
        section_id=section_id,
        text=question.text,
        question_type=question.question_type,
        order=question.order
    )
    db.add(q)
    db.commit()
    db.refresh(q)
    
    return {"id": q.id, "text": q.text, "question_type": q.question_type}
@router.delete("/{form_id}/sections/{section_id}/questions/{question_id}")
def delete_question(form_id: int, section_id: int, question_id: int, db: Session = Depends(get_db), admin = Depends(get_current_employee)):
    q = db.query(WalkaroundQuestion).filter(WalkaroundQuestion.id == question_id, WalkaroundQuestion.section_id == section_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
    db.delete(q)
    db.commit()
    return {"message": "Question deleted"}

@router.put("/{form_id}/sections/{section_id}/questions/{question_id}")
def update_question(form_id: int, section_id: int, question_id: int, question: QuestionCreate, db: Session = Depends(get_db), admin = Depends(get_current_employee)):
    q = db.query(WalkaroundQuestion).filter(WalkaroundQuestion.id == question_id, WalkaroundQuestion.section_id == section_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
    q.text = question.text
    q.question_type = question.question_type
    db.commit()
    db.refresh(q)
    return {"id": q.id, "text": q.text, "question_type": q.question_type}


@router.post("/submit")
async def submit_walkaround(
    request: Request,
    db: Session = Depends(get_db)
):
    """Walk-around form submission. Identified by badge or observe-login session."""
    import json, base64
    from datetime import datetime
    from app.models import WalkaroundSubmission, Employee, SessionRecord

    MAX_BYTES = 10 * 1024 * 1024  # 10 MB

    form = await request.form()
    try:
        form_id = int(form.get("form_id", 0))
    except (TypeError, ValueError):
        form_id = 0
    badge = (form.get("badge") or "").strip()
    responses_raw = form.get("responses", "{}")
    try:
        responses = json.loads(responses_raw)
    except Exception:
        responses = {}

    if not form_id:
        raise HTTPException(status_code=400, detail="No walk-around form selected.")

    # Identify employee
    employee = None
    if badge:
        employee = db.query(Employee).filter(Employee.badge == badge).first()
    if not employee:
        token = request.cookies.get("session_token")
        if token:
            sr = db.query(SessionRecord).filter(SessionRecord.id == token).first()
            if sr and sr.expires_at > datetime.utcnow():
                employee = sr.employee
    if not employee:
        raise HTTPException(status_code=401, detail="Please log in again.")

    photo_data = None
    video_data = None

    photo = form.get("photo")
    if photo and hasattr(photo, "filename") and photo.filename:
        content = await photo.read()
        if len(content) > MAX_BYTES:
            raise HTTPException(status_code=413, detail="Photo exceeds 10 MB limit.")
        photo_data = base64.b64encode(content).decode("ascii")

    video = form.get("video")
    if video and hasattr(video, "filename") and video.filename:
        content = await video.read()
        if len(content) > MAX_BYTES:
            raise HTTPException(status_code=413, detail="Video exceeds 10 MB limit.")
        video_data = base64.b64encode(content).decode("ascii")

    record = WalkaroundSubmission(
        employee_id=employee.id,
        form_id=form_id,
        responses=responses,
        photo_data=photo_data,
        video_data=video_data,
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    return {
        "success": True,
        "submission_id": record.id,
        "employee_name": employee.name,
        "message": "Walk-around submitted successfully."
    }

